"""Document engine - fills token-marked copies of APEX THERMOCON's own paperwork.

Two fillers, one contract::

    render(kind, payload) -> (bytes, filename)

``XlsxFiller`` and ``DocxFiller`` both take a registry entry (see
``registry.py``), replace every ``{{token}}`` in the template with a value
resolved from the payload, fill the item region, and hand back the saved
bytes.  A final sweep asserts that no ``{{`` survives anywhere - cells,
paragraphs, table cells, headers and footers.

Design rules (SOP-DESIGN §4):

* Templates are byte-for-byte copies of the reference documents with
  ``{{tokens}}`` typed into the variable cells.  Static layout is never
  touched by the engine, which is what makes fidelity provable against
  ``reference_specs/*.spec.json``.
* Swapping a template file = adjust its registry entry, zero engine changes.
* Pure Python: openpyxl + python-docx.  Nothing is ever rendered.
"""

from __future__ import annotations

import copy
import datetime as _dt
import io
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils.cell import (column_index_from_string, coordinate_from_string,
                                 range_boundaries)

from docx import Document
from docx.oxml.ns import qn

from . import dates as _dates

TEMPLATE_DIR = Path(__file__).resolve().parent / "templates"
MEDIA_DIR = TEMPLATE_DIR / "media"

TOKEN_RE = re.compile(r"\{\{([A-Za-z0-9_.]+)\}\}")


class TemplateError(RuntimeError):
    """Template and registry entry disagree, or a token was left unfilled."""


# --------------------------------------------------------------------------
# payload -> value resolution
# --------------------------------------------------------------------------

_MISSING = object()


def _dig(payload, path):
    cur = payload
    for part in path.split("."):
        if cur is None:
            return None
        if isinstance(cur, (list, tuple)):
            try:
                cur = cur[int(part)]
            except (ValueError, IndexError):
                return None
        elif isinstance(cur, dict):
            if part not in cur:
                return _MISSING
            cur = cur[part]
        else:
            cur = getattr(cur, part, _MISSING)
            if cur is _MISSING:
                return _MISSING
    return cur


def resolve(spec, payload, item=None):
    """Resolve one token spec against the payload (or an item dict)."""
    if "const" in spec:
        return spec["const"]
    source = item if item is not None else payload
    path = spec.get("path")
    value = _dig(source, path) if path else _MISSING
    if value is _MISSING or value is None:
        if "default" in spec:
            value = spec["default"]
        elif spec.get("required", False):
            raise TemplateError(f"payload is missing required field {path!r}")
        else:
            value = None
    if value is not None and "index" in spec:
        try:
            value = value[spec["index"]]
        except (IndexError, KeyError, TypeError):
            value = spec.get("default")
    render = spec.get("render")
    if render:
        fn = _dates.RENDERERS.get(render)
        if fn is None:
            raise TemplateError(f"unknown renderer {render!r}")
        value = fn(value)
    if value is None:
        value = spec.get("blank", "")
    if spec.get("upper") and isinstance(value, str):
        value = value.upper()
    return value


def build_token_map(entry, payload):
    """token name -> resolved value for every scalar token of a registry entry."""
    out = {}
    for token, spec in entry.get("tokens", {}).items():
        out[token] = resolve(spec, payload)
    return out


# --------------------------------------------------------------------------
# shared helpers
# --------------------------------------------------------------------------

def _as_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, _dt.datetime):
        return value.date().isoformat()
    if isinstance(value, _dt.date):
        return value.isoformat()
    return str(value)


def _substitute_text(text, values, consumed):
    def repl(m):
        name = m.group(1)
        if name not in values:
            raise TemplateError(f"template token {{{{{name}}}}} has no registry entry")
        consumed.add(name)
        return _as_text(values[name])
    return TOKEN_RE.sub(repl, text)


# --------------------------------------------------------------------------
# XLSX
# --------------------------------------------------------------------------

class XlsxFiller:
    """Fills an .xlsx template: scalar tokens, item slots, overflow, images."""

    def __init__(self, entry, template_dir=TEMPLATE_DIR):
        self.entry = entry
        self.path = Path(template_dir) / entry["template"]

    # -- public ------------------------------------------------------------
    def fill(self, payload) -> bytes:
        wb = load_workbook(self.path)
        ws = wb[self.entry["sheet"]] if self.entry.get("sheet") else wb.worksheets[0]

        values = build_token_map(self.entry, payload)
        self._verify_coordinates(ws)

        items_spec = self.entry.get("items")
        if items_spec:
            self._fill_items(ws, items_spec, payload)

        consumed = set()
        self._replace_everywhere(wb, values, consumed)
        self._check_consumed(values, consumed)
        self._apply_currency_numfmt(ws, payload)
        self._ensure_images(ws)
        self._sweep(wb)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # -- template integrity -------------------------------------------------
    def _verify_coordinates(self, ws):
        """Every declared coordinate really carries its declared token."""
        for token, spec in self.entry.get("tokens", {}).items():
            coord = spec.get("cell")
            if not coord:
                continue
            got = ws[coord].value
            marker = "{{%s}}" % token
            if got is None or marker not in str(got):
                raise TemplateError(
                    f"{self.entry['kind']}: template cell {coord} does not carry "
                    f"{marker} (found {got!r}) - template and registry disagree")
        items = self.entry.get("items")
        if items:
            first = items["first_row"]
            for col, spec in items["cells"].items():
                marker = "{{%s.%s}}" % (items.get("prefix", "item"), spec["path"])
                got = ws[f"{col}{first}"].value
                if got is None or marker not in str(got):
                    raise TemplateError(
                        f"{self.entry['kind']}: item slot {col}{first} does not carry "
                        f"{marker} (found {got!r})")

    # -- item slots ---------------------------------------------------------
    def _fill_items(self, ws, spec, payload):
        prefix = spec.get("prefix", "item")
        first, last = spec["first_row"], spec["last_row"]
        slots = last - first + 1
        items = payload.get(spec.get("path", "items")) or []
        if len(items) > slots:
            self._grow(ws, spec, len(items) - slots)
            last = spec["last_row"] + (len(items) - slots)

        for offset in range(last - first + 1):
            row = first + offset
            item = items[offset] if offset < len(items) else None
            for col, cell_spec in spec["cells"].items():
                cell = ws[f"{col}{row}"]
                marker = "{{%s.%s}}" % (prefix, cell_spec["path"])
                if item is None:
                    if cell.value is not None and marker in str(cell.value):
                        cell.value = None
                    continue
                value = resolve(cell_spec, payload, item=item)
                if cell.value is not None and str(cell.value).strip() == marker:
                    cell.value = None if value == "" else value
                else:                      # token embedded in surrounding text
                    cell.value = _substitute_text(str(cell.value or ""),
                                                  {f"{prefix}.{cell_spec['path']}": value},
                                                  set())

    def _grow(self, ws, spec, extra):
        """Insert ``extra`` style-cloned rows after the last slot row."""
        clone_row = spec["last_row"]
        insert_at = clone_row + 1
        inner = [str(m) for m in ws.merged_cells.ranges
                 if range_boundaries(str(m))[1] == clone_row == range_boundaries(str(m))[3]]
        height = ws.row_dimensions[clone_row].height
        insert_rows_preserving_merges(ws, insert_at, extra)

        max_col = ws.max_column
        for i in range(extra):
            target = insert_at + i
            for c in range(1, max_col + 1):
                src = ws.cell(row=clone_row, column=c)
                dst = ws.cell(row=target, column=c)
                dst._style = copy.copy(src._style)
                dst.value = None
                if src.value is not None and TOKEN_RE.search(str(src.value)):
                    dst.value = src.value
            if height is not None:
                ws.row_dimensions[target].height = height
            for merge in inner:
                lo_c, _lo_r, hi_c, _hi_r = range_boundaries(merge)
                ws.merge_cells(start_row=target, start_column=lo_c,
                               end_row=target, end_column=hi_c)

    # -- tokens -------------------------------------------------------------
    def _replace_everywhere(self, wb, values, consumed):
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not isinstance(v, str) or "{{" not in v:
                        continue
                    m = TOKEN_RE.fullmatch(v.strip())
                    if m:                          # whole-cell token: keep the native type
                        name = m.group(1)
                        if name not in values:
                            raise TemplateError(
                                f"template token {{{{{name}}}}} has no registry entry")
                        consumed.add(name)
                        resolved = values[name]
                        cell.value = None if resolved == "" else resolved
                    else:
                        cell.value = _substitute_text(v, values, consumed)
            for attr in ("oddHeader", "oddFooter", "evenHeader", "evenFooter",
                         "firstHeader", "firstFooter"):
                block = getattr(ws, attr, None)
                if block is None:
                    continue
                for part in ("left", "center", "right"):
                    piece = getattr(block, part, None)
                    if piece is not None and piece.text and "{{" in piece.text:
                        piece.text = _substitute_text(piece.text, values, consumed)

    def _check_consumed(self, values, consumed):
        unused = sorted(set(values) - consumed)
        if unused:
            raise TemplateError(
                f"{self.entry['kind']}: registry declares tokens the template never "
                f"uses: {unused}")

    def _apply_currency_numfmt(self, ws, payload):
        """Swap the template's ``$`` for the payload currency where declared.

        CONVENTIONS §9-D: the printed currency is a field, not a fixed string.
        On the ack the *header* is a token; the money **symbol** rides on the
        cell number format, so it is rewritten here (and only when the payload
        is not USD, which keeps the USD fidelity test byte-identical to the
        template).
        """
        cells = self.entry.get("currency_numfmt_cells")
        if not cells:
            return
        symbol = ((payload.get("currency") or {}).get("symbol")) or "$"
        if symbol == "$":
            return
        rows = cells if isinstance(cells, list) else []
        for coord in rows:
            if ":" in coord:
                lo_c, lo_r, hi_c, hi_r = range_boundaries(coord)
                targets = [ws.cell(row=r, column=c)
                           for r in range(lo_r, hi_r + 1)
                           for c in range(lo_c, hi_c + 1)]
            else:
                targets = [ws[coord]]
            for cell in targets:
                if '"$"' in cell.number_format:
                    cell.number_format = cell.number_format.replace('"$"', f'"{symbol}"')

    # -- images -------------------------------------------------------------
    def _ensure_images(self, ws):
        """Re-anchor declared images that openpyxl dropped on load.

        The two Test Certificates keep their logo + ISO badges in a drawing
        container openpyxl cannot read (it also holds a freeform shape), so
        ``ws._images`` comes back empty and the picture parts are lost on save.
        The template build scripts extract ``xl/media/*`` from the reference
        zip once into ``templates/media/`` and the registry declares them here;
        anything still missing after load is re-anchored as an openpyxl Image.
        """
        declared = self.entry.get("images") or []
        if not declared:
            return
        present = set()
        for img in ws._images:
            try:
                present.add((img.anchor._from.col, img.anchor._from.row))
            except Exception:
                pass
        for decl in declared:
            col_letter, row = coordinate_from_string(decl["anchor"])
            key = (column_index_from_string(col_letter) - 1, row - 1)
            if key in present:
                continue
            img = XlImage(str(MEDIA_DIR / decl["file"]))
            img.width = decl["width"]
            img.height = decl["height"]
            ws.add_image(img, decl["anchor"])

    # -- sweep --------------------------------------------------------------
    def _sweep(self, wb):
        leftovers = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and "{{" in cell.value:
                        leftovers.append(f"{ws.title}!{cell.coordinate}={cell.value!r}")
        if leftovers:
            raise TemplateError(f"{self.entry['kind']}: unfilled tokens {leftovers}")


def insert_rows_preserving_merges(ws, idx, amount):
    """``ws.insert_rows`` that actually moves merged ranges and row heights.

    openpyxl's ``insert_rows`` shifts cells only: merged ranges and row
    dimensions below the insertion point are left pointing at the old rows,
    which silently shreds every template whose footer is merged (all of ours).
    This drops the merges, inserts, then re-creates them shifted - ranges that
    *span* the insertion point are stretched rather than moved.
    """
    merges = [str(m) for m in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
    heights = {r: d.height for r, d in ws.row_dimensions.items() if d.height is not None}

    ws.insert_rows(idx, amount)

    for m in merges:
        lo_c, lo_r, hi_c, hi_r = range_boundaries(m)
        if lo_r >= idx:
            lo_r += amount
            hi_r += amount
        elif hi_r >= idx:
            hi_r += amount
        ws.merge_cells(start_row=lo_r, start_column=lo_c,
                       end_row=hi_r, end_column=hi_c)

    for r in list(ws.row_dimensions):
        ws.row_dimensions[r].height = None
    for r, h in heights.items():
        ws.row_dimensions[r + amount if r >= idx else r].height = h


# --------------------------------------------------------------------------
# DOCX
# --------------------------------------------------------------------------

class DocxFiller:
    """Fills a .docx template: tokens in paragraphs/cells, cloned item rows."""

    def __init__(self, entry, template_dir=TEMPLATE_DIR):
        self.entry = entry
        self.path = Path(template_dir) / entry["template"]

    def fill(self, payload) -> bytes:
        doc = Document(str(self.path))
        values = build_token_map(self.entry, payload)

        items_spec = self.entry.get("items")
        if items_spec:
            self._fill_item_rows(doc, items_spec, payload)

        consumed = set()
        for part in self._paragraph_sources(doc):
            _replace_in_paragraph_el(part, values, consumed)
        unused = sorted(set(values) - consumed)
        if unused:
            raise TemplateError(
                f"{self.entry['kind']}: registry declares tokens the template never "
                f"uses: {unused}")
        self._sweep(doc)

        buf = io.BytesIO()
        doc.save(buf)
        return buf.getvalue()

    # -- item rows ----------------------------------------------------------
    def _fill_item_rows(self, doc, spec, payload):
        table = doc.tables[spec["table"]]
        marker_tr = table.rows[spec["marker_row"]]._tr
        parent = marker_tr.getparent()

        rows = spec["row_builder"](payload)
        anchor = marker_tr
        for row in rows:
            new_tr = copy.deepcopy(marker_tr)
            anchor.addnext(new_tr)
            anchor = new_tr
            tcs = new_tr.findall(qn("w:tc"))
            values = {f"{spec.get('prefix', 'item')}.{k}": v
                      for k, v in row.get("cells", {}).items()}
            consumed = set()
            for tc in tcs:
                for p in tc.findall(qn("w:p")):
                    _replace_in_paragraph_el(p, values, consumed)
            for idx, mode in (row.get("vmerge") or {}).items():
                _set_vmerge(tcs[idx], mode)
        parent.remove(marker_tr)

    # -- traversal ----------------------------------------------------------
    def _paragraph_sources(self, doc):
        yield from _iter_paragraph_elements(doc.element.body)
        for section in doc.sections:
            for part in (section.header, section.footer,
                         section.even_page_header, section.even_page_footer,
                         section.first_page_header, section.first_page_footer):
                if part is None:
                    continue
                yield from _iter_paragraph_elements(part._element)

    def _sweep(self, doc):
        leftovers = []
        for p in self._paragraph_sources(doc):
            text = "".join(t.text or "" for t in p.iter(qn("w:t")))
            if "{{" in text:
                leftovers.append(text[:120])
        if leftovers:
            raise TemplateError(f"{self.entry['kind']}: unfilled tokens {leftovers}")


def _iter_paragraph_elements(root):
    for p in root.iter(qn("w:p")):
        yield p


def _replace_in_paragraph_el(p_el, values, consumed):
    """Replace ``{{token}}`` inside one ``<w:p>``.

    Templates author every token as a single run, so the fast path is a plain
    per-run substitution.  The slow path exists because Word will happily split
    a run on a spell-check boundary after a human edit: runs that jointly hold
    a token get merged into the first of them (which keeps its formatting)
    before substituting.
    """
    runs = p_el.findall(qn("w:r"))
    if not runs:
        return
    for r in runs:
        for t in r.findall(qn("w:t")):
            if t.text and "{{" in t.text:
                new = _substitute_text(t.text, values, consumed)
                t.text = new
                if new != new.strip() or new.startswith(" ") or new.endswith(" "):
                    t.set(qn("xml:space"), "preserve")

    text = "".join((t.text or "") for r in p_el.findall(qn("w:r"))
                   for t in r.findall(qn("w:t")))
    if "{{" not in text:
        return

    # slow path - a token straddles run boundaries
    runs = p_el.findall(qn("w:r"))
    spans, pos = [], 0
    for r in runs:
        ts = r.findall(qn("w:t"))
        length = sum(len(t.text or "") for t in ts)
        spans.append((pos, pos + length, r))
        pos += length
    for m in reversed(list(TOKEN_RE.finditer(text))):
        touched = [(s, e, r) for (s, e, r) in spans if s < m.end() and e > m.start()]
        if not touched:
            continue
        first = touched[0]
        merged = text[first[0]:touched[-1][1]]
        merged = _substitute_text(merged, values, consumed)
        ts = first[2].findall(qn("w:t"))
        if not ts:
            continue
        ts[0].text = merged
        ts[0].set(qn("xml:space"), "preserve")
        for extra in ts[1:]:
            first[2].remove(extra)
        for _s, _e, r in touched[1:]:
            r.getparent().remove(r)
        text = text[:first[0]] + merged + text[touched[-1][1]:]
        spans = []
        pos = 0
        for r in p_el.findall(qn("w:r")):
            length = sum(len(t.text or "") for t in r.findall(qn("w:t")))
            spans.append((pos, pos + length, r))
            pos += length


def _set_vmerge(tc, mode):
    """``mode`` is ``'restart'``, ``'continue'`` or ``None`` (no vertical merge)."""
    tcPr = tc.get_or_add_tcPr()
    tcPr._remove_vMerge()
    if mode is None:
        return
    vmerge = tcPr.get_or_add_vMerge()
    if mode == "restart":
        vmerge.set(qn("w:val"), "restart")
    else:
        if vmerge.get(qn("w:val")) is not None:
            del vmerge.attrib[qn("w:val")]


# --------------------------------------------------------------------------
# public entry point
# --------------------------------------------------------------------------

def filler_for(entry, template_dir=TEMPLATE_DIR):
    if entry["format"] == "xlsx":
        return XlsxFiller(entry, template_dir)
    if entry["format"] == "docx":
        return DocxFiller(entry, template_dir)
    raise TemplateError(f"unknown template format {entry['format']!r}")


def render(kind, payload, template_dir=TEMPLATE_DIR):
    """Render one paper.  Returns ``(bytes, download_filename)``."""
    from . import registry
    entry = registry.entry(kind)
    prepared = registry.prepare(kind, payload)
    data = filler_for(entry, template_dir).fill(prepared)
    return data, registry.filename(kind, payload)
