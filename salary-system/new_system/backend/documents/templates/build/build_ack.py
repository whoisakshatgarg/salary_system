"""ack.xlsx  <-  "PO Acknowledgement Template (converted).xlsx".

Provenance: the blank house acknowledgement, converted from BIFF .xls by the
Phase-0 pure-Python ``tools/xls2xlsx.py``.  Only three things change:

1. variable cells get ``{{tokens}}`` (styles untouched),
2. CONVENTIONS §8 typo fix ``Payment. Terms`` -> ``Payment Terms`` (E15),
3. print setup is set explicitly - BIFF reading loses it, so the converted
   workbook has none at all (documented deviation).
"""

from __future__ import annotations

from openpyxl import load_workbook

from . import common

SHEET = "01"

CELL_TOKENS = {
    # Bill To / Ship To blocks (5 line cells each)
    "A7": "bill_to_1", "A8": "bill_to_2", "A9": "bill_to_3",
    "A10": "bill_to_4", "A11": "bill_to_5",
    "G7": "ship_to_1", "G8": "ship_to_2", "G9": "ship_to_3",
    "G10": "ship_to_4", "G11": "ship_to_5",
    # right-hand header grid
    "F6": "cust_po",
    "F7": "po_date",
    "F9": "quotation_ref",
    "C12": "client_code",
    "F12": "ack_ref",
    "F13": "ack_date",
    "F14": "price_basis",
    "F15": "payment_terms",
    "F16": "ship_date",
    "F17": "wo_no_long",
    # CONTACTS
    "B14": "contact_name",
    "B15": "contact_email",
    "B16": "contact_tel",
    "B17": "contact_fax",
    # item grid header + total
    "I18": "currency_header",
    "I38": "total",
    # remittance block
    "A44": "remit_intro",
    "A45": "remittance_block",
    "A48": "beneficiary_label",
    "A49": "beneficiary_address",
    "A50": "beneficiary_account",
}

ITEM_COLUMNS = {
    "A": "sno", "B": "code", "C": "description", "F": "material",
    "G": "qty", "H": "unit", "I": "unit_price", "J": "total",
}
ITEMS_FIRST, ITEMS_LAST = 20, 37


def build():
    src = common.sample("PO Acknowledgement Template (converted).xlsx")
    wb = load_workbook(src)
    ws = wb[SHEET]

    # (2) CONVENTIONS §8 typo fix
    assert ws["E15"].value == "Payment. Terms", ws["E15"].value
    ws["E15"] = "Payment Terms"

    # (1) tokens
    common.tokenise(ws, CELL_TOKENS)
    common.fill_slots(ws, ITEMS_FIRST, ITEMS_LAST, ITEM_COLUMNS)

    # (3) print setup lost in BIFF conversion
    common.set_print_setup(ws, "portrait", "A4", 1, 0)

    dst = common.out("ack.xlsx")
    wb.save(dst)
    return dst


if __name__ == "__main__":
    print(build())
