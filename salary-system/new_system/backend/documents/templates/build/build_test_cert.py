"""test_cert.xlsx  <-  "Test Certiticate PO59812-EI-047.xlsx".

Provenance: the material test certificate, native .xlsx, landscape.  Changes:

1. variable cells get ``{{tokens}}``,
2. CONVENTIONS §8 typo fix ``nas been tested`` -> ``has been tested`` (A33),
3. the logo + ISO-badge pictures are lifted out of the reference zip into
   ``templates/media/`` and re-anchored as openpyxl Image objects, because the
   reference keeps them in a drawing container that also holds a freeform
   shape - openpyxl reads none of it and would drop both pictures on save.

The five spare chemistry columns P..T keep their ``-`` default; the engine
writes an element symbol into the header when the payload carries
``extra_elements``.
"""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage

from . import common

SHEET = "Sheet2"

CELL_TOKENS = {
    "Q3": "cert_no",
    "Q4": "cert_date",          # real date, template numfmt mm-dd-yy
    "C5": "customer_line",
    "C6": "po",
    "F6": "po_date",            # real date, mm-dd-yy
    "K6": "invoice_no",
    "Q6": "invoice_date",       # real date, mm-dd-yy
    # spare chemistry column headers (default '-')
    "P8": "spare_head_1", "Q8": "spare_head_2", "R8": "spare_head_3",
    "S8": "spare_head_4", "T8": "spare_head_5",
}

ITEM_COLUMNS = {
    "A": "sno", "B": "item", "C": "size", "D": "qty",
    "E": "component", "F": "heat_no", "G": "material",
    "H": "chem.C", "I": "chem.Mn", "J": "chem.Si", "K": "chem.P",
    "L": "chem.S", "M": "chem.Cr", "N": "chem.Ni", "O": "chem.Mo",
    "P": "spare_1", "Q": "spare_2", "R": "spare_3",
    "S": "spare_4", "T": "spare_5",
}
ITEMS_FIRST, ITEMS_LAST = 9, 28

CERTIFY_CELL = "A33"
CERTIFY_TYPO = ("We hereby Certify that the material described above nas been tested by us "
                "and complies to the material requirements of the Purchase Order.")
CERTIFY_FIXED = CERTIFY_TYPO.replace("nas been tested", "has been tested")

# Display sizes read from xl/drawings/drawing1.xml of the reference and
# converted at 9525 EMU/px, so the regenerated anchors carry the *identical*
# EMU extents (973455x842010 and 933450x476250).
IMAGES = [
    {"file": "apex_logo.png", "anchor": "A1", "width": 102.2, "height": 88.4},
    {"file": "apex_iso.png", "anchor": "K1", "width": 98.0, "height": 50.0},
]


def build():
    src = common.sample("Test Certiticate PO59812-EI-047.xlsx")
    common.extract_media(src, {"image1.png": "apex_logo.png",
                               "image2.png": "apex_iso.png"})

    wb = load_workbook(src)
    ws = wb[SHEET]

    assert ws[CERTIFY_CELL].value == CERTIFY_TYPO, ws[CERTIFY_CELL].value
    ws[CERTIFY_CELL] = CERTIFY_FIXED

    common.tokenise(ws, CELL_TOKENS)
    common.fill_slots(ws, ITEMS_FIRST, ITEMS_LAST, ITEM_COLUMNS)

    assert not ws._images, "reference unexpectedly exposed its pictures"
    for decl in IMAGES:
        img = XlImage(str(common.MEDIA_DIR / decl["file"]))
        img.width, img.height = decl["width"], decl["height"]
        ws.add_image(img, decl["anchor"])

    dst = common.out("test_cert.xlsx")
    wb.save(dst)
    return dst


if __name__ == "__main__":
    print(build())
