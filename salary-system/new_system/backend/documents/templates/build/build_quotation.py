"""quotation.xlsx  -  rebuilt from the measured geometry of the reference PDF.

Provenance: ``sample docs/Thermosense-Quotation-316.pdf`` is the only
quotation reference the company has (no editable original survives), so there
is no ``reference_specs`` entry.  Every ruling line and every word position
was measured off the PDF in points
(``phase0/dumps/quotation_geometry.txt``, page 595.32 x 841.92 = A4 portrait)
and is reproduced here as an Excel grid.

**Pending owner visual sign-off.**

Geometry
--------
Content frame ``x = 38.9 .. 522.0 pt``.  Item-table column boundaries::

    38.9 | 58.0 | 107.2 | 357.3 | 393.1 | 420.0 | 465.1 | 520.4
       A     B       C       D       E       F       G

points -> Excel character widths.  Excel stores a column width in characters
of the workbook's *normal* font and renders it as
``px = round(chars * MDW + 5)``, where MDW is the maximum digit width in
pixels (7 px for the Calibri 11 normal font openpyxl writes) and 1 pt =
96/72 px.  Inverting that, and keeping the 5 px cell padding out of the
proportion so the column *boundaries* land where the PDF has them::

    chars = (pt * 96/72 - 5) / 7

which works out to an effective **5.55 pt per character** across this sheet
(the brief's ~5.4 estimate); the seven columns sum to 86.71 chars = 642 px =
481.5 pt, exactly the measured frame width.

Row heights are the measured y-deltas of the PDF's text lines and rules, in
points (Excel's row-height unit is also points), with the 30 body slots
sharing the measured 255.3 -> 585.3 band evenly.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from . import common

# --- geometry ---------------------------------------------------------------
PT_PER_PX = 72 / 96
MDW_PX = 7          # Calibri 11 normal-font max digit width, as Excel measures it


def pt_to_chars(pt: float) -> float:
    return round((pt / PT_PER_PX - 5) / MDW_PX, 2)


COL_BOUNDS = [38.9, 58.0, 107.2, 357.3, 393.1, 420.0, 465.1, 520.4]
COL_WIDTHS_PT = [COL_BOUNDS[i + 1] - COL_BOUNDS[i] for i in range(7)]

ROW_HEIGHTS = {
    1: 14.2, 2: 14.0, 3: 10.5, 4: 10.5, 5: 9.9,          # letterhead
    6: 10.4, 7: 10.4, 8: 10.4, 9: 10.4,                  # client / RFQ block
    10: 10.4,                                            # client code
    11: 10.4, 12: 10.4, 13: 10.4, 14: 10.5,              # attn / tel / fax / e-mail
    15: 10.0, 16: 9.8,                                   # intro sentence (merged)
    17: 10.5, 18: 9.6,                                   # item table header
    49: 10.7,                                            # validity line
    50: 13.4,                                            # total quotation value
    51: 10.4, 52: 10.4, 53: 10.4, 54: 10.4, 55: 10.5, 56: 10.4,
    57: 10.4, 58: 10.0, 59: 9.3, 60: 9.4,
}
ITEMS_FIRST, ITEMS_LAST = 19, 48
ITEM_ROW_HEIGHT = 10.9
LAST_ROW = 60

# --- type -------------------------------------------------------------------
BLUE = "FF1F497D"
F_TITLE = dict(name="Arial", size=11.5, bold=True)
F_BODY = dict(name="Arial", size=8)
F_ITEM = dict(name="Arial", size=7.5)

THIN = Side(style="thin")
MEDIUM = Side(style="medium")

# CONVENTIONS §5: Indian digit grouping on quotation quantities.
INDIAN_QTY = r"[>=10000000]##\,##\,##\,##0;[>=100000]##\,##\,##0;#,##0"
MONEY_3 = '"$"#,##0.000'
MONEY_2 = '"$"#,##0.00'

STATIC = {
    "A1": "QUOTATION",
    "A2": "APEX THERMOCON Pvt. Ltd.",
    "A3": "A-2/15, Sector 17, Kavi Nagar Industrial Area , Ghaziabad 201 002 , UP , INDIA",
    "A4": ("Tel. No. 91-120- 4167651 / 91-9810127235    Fax No. 91- 120- 4167561  ~  "
           "E-mail : sales@apexthermocon.com"),
    "A5": "www.apexthermocon.com",
    "D6": "R.F.Q.Ref.No", "D7": "R.F.Q.Date", "D8": "Quotation Dt.", "D9": "Quotation No",
    "A10": "Client Code",
    "A11": "Attn.", "A12": "Tel. No", "A13": "Fax No.", "A14": "E-Mail",
    "A17": "Sr\nNo.", "B17": "Product\nCode", "C17": "Description",
    "D17": "Qty", "E17": "Unit", "F18": "Unit", "G18": "Total",
    "A50": "Total Quotation Value",          # CONVENTIONS §8: PDF says 'Vaue'
    "A51": "Price basis", "A52": "Lead Time", "A53": "Payment Terms",
    "A54": "Guarantee", "A55": "Taxes & Duties", "A56": "Note",
    "A57": "For Apex Thermocon Pvt. Ltd.",
    "D60": "Format : AT/QTN/EXP/01",
}

MERGES = [
    "A1:G1", "A2:G2", "A3:G3", "A4:G4", "A5:G5",
    "A6:C6", "A7:C7", "A8:C8", "A9:C9",
    "D6:E6", "D7:E7", "D8:E8", "D9:E9",
    "F6:G6", "F7:G7", "F8:G8", "F9:G9",
    # the x=357.3 rule runs through the Client Code row too, so the value cell
    # is C alone and D..G is the (empty) right-hand block, as on rows 11-14
    "A10:B10", "D10:G10",
    "A11:B11", "A12:B12", "A13:B13", "A14:B14",
    "D11:G11", "D12:G12", "D13:G13", "D14:G14",
    "A15:G16",
    "A17:A18", "B17:B18", "C17:C18", "D17:D18", "E17:E18", "F17:G17",
    "A50:F50",
    "C51:G51", "C52:G52", "C53:G53", "C54:G54", "C55:G55", "C56:G56",
    "A51:B51", "A52:B52", "A53:B53", "A54:B54", "A55:B55", "A56:B56",
    "A57:G57", "A59:C59", "D59:G59", "A60:C60", "D60:G60",
]

CELL_TOKENS = {
    "A6": "customer_name", "A7": "customer_line_2", "A8": "customer_line_3",
    "A9": "customer_country",
    "F6": "rfq_ref", "F7": "rfq_date", "F8": "quotation_date", "F9": "number",
    "C10": "client_code",
    "C11": "attn", "C12": "tel", "C13": "fax", "C14": "email",
    "A15": "intro",
    "F17": "currency_header",
    "C49": "validity_line",
    "G50": "total",
    "C51": "price_basis", "C52": "lead_time", "C53": "payment_terms",
    "C54": "guarantee", "C55": "taxes_duties", "C56": "note",
}

ITEM_COLUMNS = {
    "A": "sno", "B": "code", "C": "description", "D": "qty",
    "E": "unit", "F": "unit_price", "G": "total",
}


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    for i, wpt in enumerate(COL_WIDTHS_PT):
        ws.column_dimensions[get_column_letter(i + 1)].width = pt_to_chars(wpt)
    for row, h in ROW_HEIGHTS.items():
        ws.row_dimensions[row].height = h
    for row in range(ITEMS_FIRST, ITEMS_LAST + 1):
        ws.row_dimensions[row].height = ITEM_ROW_HEIGHT

    for rng in MERGES:
        ws.merge_cells(rng)

    # ---- static type -------------------------------------------------------
    for coord, text in STATIC.items():
        ws[coord] = text

    ws["A1"].font = Font(underline="single", **F_TITLE)
    ws["A2"].font = Font(color=BLUE, **F_TITLE)
    for coord in ("A3", "A4"):
        ws[coord].font = Font(italic=True, bold=True, name="Arial", size=8)
    ws["A5"].font = Font(underline="single", name="Arial", size=8)
    for coord in ("A1", "A2", "A3", "A4", "A5"):
        ws[coord].alignment = Alignment(horizontal="center", vertical="center")

    bold_labels = ["D6", "D7", "D8", "D9", "A10", "A11", "A12", "A13", "A14",
                   "A17", "B17", "C17", "D17", "E17", "F18", "G18", "A50",
                   "A51", "A52", "A53", "A54", "A55", "A56", "A57", "D60"]
    for coord in bold_labels:
        ws[coord].font = Font(bold=True, **F_BODY)
    ws["D60"].font = Font(**F_BODY)                      # format strip is not bold
    for coord in ("A17", "B17", "C17", "D17", "E17", "F18", "G18"):
        ws[coord].alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
    for coord in ("A50",):
        ws[coord].alignment = Alignment(horizontal="left", vertical="center")

    # ---- tokens ------------------------------------------------------------
    common.tokenise(ws, CELL_TOKENS)
    for coord in CELL_TOKENS:
        ws[coord].font = Font(**F_BODY)
        ws[coord].alignment = Alignment(horizontal="left", vertical="center")
    ws["F17"].font = Font(bold=True, **F_BODY)
    ws["F17"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C14"].font = Font(underline="single", **F_BODY)
    ws["A15"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws["C49"].font = Font(bold=True, **F_ITEM)
    ws["C49"].alignment = Alignment(horizontal="center", vertical="center")
    ws["G50"].font = Font(bold=True, **F_BODY)
    ws["G50"].alignment = Alignment(horizontal="right", vertical="center")
    ws["G50"].number_format = MONEY_2

    ws["A59"] = "Approved by : {{approved_by}}"
    ws["A59"].font = Font(bold=True, **F_BODY)
    ws["D59"] = "Prepared by {{prepared_by}}"
    ws["D59"].font = Font(bold=True, **F_BODY)

    # ---- item slots --------------------------------------------------------
    common.fill_slots(ws, ITEMS_FIRST, ITEMS_LAST, ITEM_COLUMNS)
    for row in range(ITEMS_FIRST, ITEMS_LAST + 1):
        for col, align in (("A", "center"), ("B", "left"), ("C", "left"),
                           ("D", "right"), ("E", "center"),
                           ("F", "right"), ("G", "right")):
            cell = ws[f"{col}{row}"]
            cell.font = Font(**F_ITEM)
            cell.alignment = Alignment(horizontal=align, vertical="center")
        ws[f"D{row}"].number_format = INDIAN_QTY
        ws[f"F{row}"].number_format = MONEY_3
        ws[f"G{row}"].number_format = MONEY_2

    _draw_rules(ws)
    common.set_print_setup(ws, "portrait", "A4", 1, 0)
    ws.page_margins.left = 0.52
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3

    dst = common.out("quotation.xlsx")
    wb.save(dst)
    return dst


# --- ruling lines -----------------------------------------------------------

def _draw_rules(ws):
    plan = {}

    def edge(row, col, side, style):
        plan.setdefault((row, col), {})[side] = style

    def hline(row, c1, c2, style):                     # bottom rule under `row`
        for c in range(c1, c2 + 1):
            edge(row, c, "bottom", style)

    def vline(col, r1, r2, style):                     # rule on the right of `col`
        for r in range(r1, r2 + 1):
            edge(r, col, "right", style)

    # outer frame (x 37.3/522.0, y 62.2/711.9)
    for r in range(1, LAST_ROW + 1):
        edge(r, 1, "left", MEDIUM)
        edge(r, 7, "right", MEDIUM)
    for c in range(1, 8):
        edge(1, c, "top", MEDIUM)
        edge(LAST_ROW, c, "bottom", MEDIUM)

    hline(5, 1, 7, THIN)          # y 119.8  under the www line
    hline(9, 1, 3, THIN)          # y 161.5  under the client block only (38.9..358.0)
    hline(10, 1, 7, THIN)         # y 172.0  under the Client Code row
    hline(14, 1, 7, THIN)         # y 213.7  under the contacts block
    hline(16, 1, 7, THIN)         # y 233.7  above the item table
    hline(17, 6, 7, THIN)         # y 244.1  under 'Prices (...)' only
    hline(18, 1, 7, THIN)         # y 254.5  under the item-table header
    hline(49, 1, 7, THIN)         # y 593.0  under the validity line
    hline(50, 1, 7, MEDIUM)       # y 606.4  under Total Quotation Value
    hline(59, 1, 7, THIN)         # y 700.3  above the format strip

    vline(3, 6, 14, THIN)         # x 357.3, y 120.5..214.5
    vline(5, 6, 10, THIN)         # x 420.0, y 120.5..172.7
    vline(2, 10, 14, THIN)        # x 107.2, y 162.3..214.5
    vline(1, 17, 49, THIN)        # x  58.0, y 234.4..593.7
    vline(2, 19, 49, THIN)        # x 107.2, y 255.3..593.7
    vline(3, 17, 49, THIN)        # x 357.3, y 234.4..593.7
    vline(4, 17, 49, THIN)        # x 393.1, y 234.4..593.7
    vline(5, 17, 49, THIN)        # x 420.0, y 234.4..593.7
    vline(6, 18, 49, THIN)        # x 465.1, y 244.8..593.0
    vline(6, 50, 50, MEDIUM)      # x 464.7, y 593.0..608.0
    vline(3, 60, 60, THIN)        # x 357.3, y 701.0..710.4

    blank = Side()
    for (row, col), sides in plan.items():
        cell = ws.cell(row=row, column=col)
        cell.border = Border(left=sides.get("left", blank),
                             right=sides.get("right", blank),
                             top=sides.get("top", blank),
                             bottom=sides.get("bottom", blank))


if __name__ == "__main__":
    print(build())
