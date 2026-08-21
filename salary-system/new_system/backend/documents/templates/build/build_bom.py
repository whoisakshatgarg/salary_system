"""bom.xlsx  <-  "Apex_BOM_Template_PLACEHOLDER.xlsx".

Provenance: our own placeholder BOM (CONVENTIONS §9-G) - there is no official
company format yet, so the file keeps its grey PLACEHOLDER footnote at A40
**verbatim** until the owner supplies one.  Changes:

1. variable cells get ``{{tokens}}``,
2. the two grey italic example rows lose their values AND their example
   styling: rows 9/10 are normalised to the black, non-italic font of slot
   rows 11-32, because real BOM lines must not print grey (owner, 2026-08-21).
   Font only - borders, alignment, wrap and number formats are untouched,
3. nothing else.  Its two pictures are ordinary openpyxl images and
   round-trip; they are declared in the registry so a future openpyxl
   regression is caught rather than silently dropping them.
"""

from __future__ import annotations

from copy import copy

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from . import common

SHEET = "BOM"

CELL_TOKENS = {
    "Q3": "bom_no",
    "Q4": "bom_date",
    "C5": "customer_line",
    "C6": "po",
    "F6": "po_date",
    "K6": "wo_no",
    "Q6": "part_assy",
}

ITEM_COLUMNS = {
    "A": "sno", "B": "part_no", "C": "description", "F": "size",
    "G": "material", "H": "heat_or_os", "J": "source",
    "L": "qty_per", "N": "total_qty", "P": "unit", "Q": "remarks",
}
ITEMS_FIRST, ITEMS_LAST = 9, 32

FOOTNOTE = ("PLACEHOLDER (proposed format code AT/BOM/EXP/01) - replace with the official "
            "BOM format when available. Grey italic rows are examples; overwrite them. "
            "Outsourced items carry an OS ID in place of a Heat No.")


EXAMPLE_ROWS = (9, 10)
STYLE_ROW = 11                      # the first plain slot row
GREY = "FF808080"


def _normalise_example_rows(ws):
    """Give the two example rows the plain slot font of rows 11-32.

    The placeholder marked its sample lines grey italic; real BOM lines must
    not print grey (owner, 2026-08-21).  Only the *font* is replaced - borders,
    alignment, wrap and number formats stay exactly as the reference has them.
    """
    changed = []
    for row in EXAMPLE_ROWS:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            font = cell.font
            grey_italic = bool(font.italic) and getattr(font.color, "rgb", None) == GREY
            if not grey_italic:
                continue                       # spacer columns are already plain
            cell.font = copy(ws.cell(row=STYLE_ROW, column=col).font)
            changed.append(f"{get_column_letter(col)}{row}")
    assert len(changed) == 22, changed          # 11 value columns x 2 rows
    return changed


def build():
    src = common.sample("Apex_BOM_Template_PLACEHOLDER.xlsx")
    common.extract_media(src, {"image1.png": "apex_logo.png",
                               "image2.png": "apex_iso.png"})

    wb = load_workbook(src)
    ws = wb[SHEET]

    assert ws["A40"].value == FOOTNOTE, ws["A40"].value          # kept verbatim
    assert len(ws._images) == 2, ws._images

    _normalise_example_rows(ws)
    common.tokenise(ws, CELL_TOKENS)
    common.fill_slots(ws, ITEMS_FIRST, ITEMS_LAST, ITEM_COLUMNS)

    dst = common.out("bom.xlsx")
    wb.save(dst)
    return dst


if __name__ == "__main__":
    print(build())
