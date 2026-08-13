"""Excel salary-sheet exporters — the two formats APEX THERMOCON already uses.

Faithful ports of the legacy ``convertCEO.py`` (20-column reconciliation sheet
for the CEO) and ``convertSalDist.py`` (16-column distribution slip), preserving
the exact layout: a merged bold title row, a bold header band, then a
thick-bordered two-row block per employee with the ``+ = -`` advance-ledger and
salary-breakdown cells, plus the same per-column widths.

Unlike the old scripts, numbers come from the clean ``pay`` table by name (not by
fragile tuple position), so the off-by-one and round-before-divide bugs are gone.
Hand a real sample .xlsx over and these can be tuned to byte-match.
"""

from __future__ import annotations

import calendar
import io
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from .engine import days_in_period

CENTER = Alignment(horizontal="center")
LEFT = Alignment(horizontal="left")
BOLD = Font(bold=True)


def _thick_border(ws, r1, c1, r2, c2):
    thick = Side(border_style="thick", color="000000")
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cell in row:
            top = thick if cell.row == r1 else None
            bottom = thick if cell.row == r2 else None
            left = thick if cell.column == c1 else None
            right = thick if cell.column == c2 else None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def _set_widths(ws, widths: dict[str, int]):
    # Iterate by index (not ws.columns) so merged title cells, which lack
    # .column_letter, don't blow up width auto-sizing.
    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        if letter in widths:
            ws.column_dimensions[letter].width = widths[letter]
            continue
        longest = max(
            (len(str(c.value)) for c in ws[letter] if c.value is not None), default=0
        )
        ws.column_dimensions[letter].width = longest + 2


def _period_label(period: str) -> tuple[str, str, int, str]:
    year, month = (int(x) for x in period.split("-"))
    month_name = calendar.month_name[month]
    tdays = days_in_period(year, month)
    title = f"{month_name}-{year}"
    return month_name, title, tdays, date.today().strftime("%d-%m-%Y")


def _gather(conn, period: str):
    """Per-employee export rows: the published `pay` record (the source of truth,
    including any CEO attendance overrides) + employee + period advance totals."""
    rows = conn.execute(
        """
        SELECT p.*, e.name AS emp_name, e.dept, e.rem_advance,
               COALESCE((SELECT SUM(amount) FROM advance a
                         WHERE a.employee_id=p.employee_id AND a.type='CR'
                         AND substr(a.txn_date,1,7)=p.period), 0) AS taken_adv
        FROM pay p
        JOIN employee e ON e.id = p.employee_id
        WHERE p.period = ?
        ORDER BY p.employee_id
        """,
        (period,),
    ).fetchall()
    return [dict(r) for r in rows]


# --------------------------------------------------------------------------- #
# CEO sheet — 20 columns (A..T)
# --------------------------------------------------------------------------- #
def build_ceo(conn, period: str) -> tuple[bytes, str]:
    month_name, title, tdays, today = _period_label(period)
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=2, column=1, value=f"Month:\t{month_name}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)
    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=9)
    ws.cell(row=2, column=6, value="Paid On").alignment = CENTER
    ws.cell(row=2, column=8, value=today).alignment = CENTER
    ws.cell(row=2, column=11, value="Days").alignment = CENTER
    ws.cell(row=2, column=12, value=tdays)
    for col, label in [(15, "Total"), (16, "Chq."), (17, "Cash"),
                       (18, "Days"), (19, "Hrs."), (20, "T")]:
        ws.cell(row=2, column=col, value=label).alignment = CENTER
    _thick_border(ws, 1, 1, 2, 20)

    r = 3
    for it in _gather(conn, period):
        rem = it["rem_advance"]
        taken = int(it["taken_adv"] or 0)
        adj = it["adv_deducted"]
        att_pct = it["attendance_percentage"] or 0
        att_days = round((att_pct / 100) * tdays)

        # Line 1: id, name, advance ledger, attendance block, bonus.
        ws.cell(row=r, column=1, value=it["employee_id"]).alignment = CENTER
        c = ws.cell(row=r, column=2, value=it["emp_name"]); c.alignment = LEFT; c.font = BOLD
        ws.cell(row=r, column=3, value=rem - taken + adj)
        ws.cell(row=r, column=4, value="+")
        ws.cell(row=r, column=5, value=taken)
        ws.cell(row=r, column=6, value="=")
        ws.cell(row=r, column=7, value=rem + adj)
        ws.cell(row=r, column=8, value="-")
        ws.cell(row=r, column=9, value=adj)
        ws.cell(row=r, column=10, value="=")
        ws.cell(row=r, column=11, value=rem)
        bonus_cell = ws.cell(
            row=r, column=15,
            value=f"Bonus: {it['bonus']}" if it["bonus_status"] == "Y" else "Bonus: 0",
        )
        bonus_cell.alignment = LEFT
        ws.cell(row=r, column=18, value=att_days)
        ws.cell(row=r, column=19, value=it["overtime_hours"] or 0)
        ws.cell(row=r, column=20, value=round(it["refreshment_pay"]))

        # Line 2: salary breakdown  base_att + ot + refresh - (pf+esi) - adj = total
        ws.cell(row=r + 1, column=2, value=it["base"]).font = BOLD
        ws.cell(row=r + 1, column=3, value=round(it["base_att"]))
        ws.cell(row=r + 1, column=4, value="+")
        ws.cell(row=r + 1, column=5, value=round(it["overtime_pay"]))
        ws.cell(row=r + 1, column=6, value="+")
        ws.cell(row=r + 1, column=7, value=round(it["refreshment_pay"]))
        ws.cell(row=r + 1, column=8, value="=")
        gross = round(it["base_att"] + it["overtime_pay"] + it["refreshment_pay"])
        ws.cell(row=r + 1, column=9, value=gross).font = BOLD
        ws.cell(row=r + 1, column=10, value="-")
        ws.cell(row=r + 1, column=11, value=it["pf"] + it["esi"])
        ws.cell(row=r + 1, column=12, value="-")
        ws.cell(row=r + 1, column=13, value=adj)
        ws.cell(row=r + 1, column=14, value="=")
        ws.cell(row=r + 1, column=15, value=it["total"]).font = BOLD
        ws.cell(row=r + 1, column=16, value=it["cheque"])
        ws.cell(row=r + 1, column=17, value=it["cash"])

        _thick_border(ws, r, 1, r + 1, 20)
        r += 2

    # Title row, bold header band.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
    tcell = ws.cell(row=1, column=1, value=f"Salary Sheet:   {title}")
    tcell.alignment = CENTER
    tcell.font = BOLD
    for cell in ws["A2:S2"][0]:
        cell.font = BOLD
    _set_widths(ws, {"A": 3, "B": 20, "M": 8, "N": 3, "O": 10, "P": 8,
                     "Q": 8, "R": 5, "S": 5, "T": 4, "D": 5, "H": 5, "F": 5, "L": 3})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"Salary_{title}_CEO1.xlsx"


# --------------------------------------------------------------------------- #
# Distribution slip — 16 columns (A..P)
# --------------------------------------------------------------------------- #
def build_distribution(conn, period: str) -> tuple[bytes, str]:
    month_name, title, tdays, today = _period_label(period)
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=2, column=1, value=f"Month:\t\t{month_name}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)
    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=9)
    ws.cell(row=2, column=6, value="Paid On").alignment = CENTER
    ws.cell(row=2, column=8, value=today).alignment = CENTER
    ws.cell(row=2, column=11, value=tdays).alignment = CENTER
    for col, label in [(12, "Chq."), (13, "Cash"), (14, "Days"), (15, "Hrs."), (16, "T")]:
        ws.cell(row=2, column=col, value=label).alignment = CENTER
    _thick_border(ws, 1, 1, 2, 16)

    r = 3
    for it in _gather(conn, period):
        rem = it["rem_advance"]
        taken = int(it["taken_adv"] or 0)
        adj = it["adv_deducted"]
        att_pct = it["attendance_percentage"] or 0
        att_days = round((att_pct / 100) * tdays)
        deductions = it["pf"] + it["esi"]

        ws.cell(row=r, column=1, value=it["employee_id"]).alignment = CENTER
        c = ws.cell(row=r, column=2, value=it["emp_name"]); c.alignment = LEFT; c.font = BOLD
        ws.cell(row=r, column=3, value=rem - taken + adj)
        ws.cell(row=r, column=4, value="+")
        ws.cell(row=r, column=5, value=taken)
        ws.cell(row=r, column=6, value="=")
        ws.cell(row=r, column=7, value=rem + adj)
        ws.cell(row=r, column=8, value="-")
        ws.cell(row=r, column=9, value=adj)
        ws.cell(row=r, column=10, value="=")
        ws.cell(row=r, column=11, value=rem)
        ws.cell(row=r, column=14, value=att_days)
        ws.cell(row=r, column=15, value=it["overtime_hours"] or 0)
        ws.cell(row=r, column=16, value=round(it["refreshment_pay"]))

        # Line 2: total + deductions + adj = gross ; - deductions - adj = total
        ws.cell(row=r + 1, column=3, value=it["total"] + deductions + adj)
        ws.cell(row=r + 1, column=4, value="-")
        ws.cell(row=r + 1, column=5, value=deductions)
        ws.cell(row=r + 1, column=6, value="-")
        ws.cell(row=r + 1, column=7, value=adj)
        ws.cell(row=r + 1, column=8, value="=")
        ws.cell(row=r + 1, column=9, value=it["total"]).font = BOLD
        ws.cell(row=r + 1, column=12, value=it["cheque"])
        ws.cell(row=r + 1, column=13, value=it["cash"])

        _thick_border(ws, r, 1, r + 1, 16)
        r += 2

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    tcell = ws.cell(row=1, column=1, value=f"Salary Slip:   {title}")
    tcell.alignment = CENTER
    tcell.font = BOLD
    for cell in ws["A2:P2"][0]:
        cell.font = BOLD
    _set_widths(ws, {"A": 3, "B": 18, "M": 10, "N": 5, "O": 5, "P": 4,
                     "D": 3, "H": 3, "F": 3, "J": 3, "L": 10})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"Salary_{title}_SalaryDist.xlsx"
