#!/usr/bin/env python3
"""One-time faithful .xls -> .xlsx conversion, pure Python (xlrd + openpyxl).

The two PO Acknowledgement references are legacy BIFF .xls files; everything
downstream (format_spec.py, the document engine) speaks .xlsx. This rebuilds
the workbook cell by cell from what xlrd exposes: values (incl. dates via the
book's datemode), merged ranges, column widths, row heights, fonts, borders,
alignment and number formats.

Not carried over (xlrd does not expose them): print setup / print areas,
embedded images, cell fills beyond pattern foregrounds. The acknowledgement
references have none of these except print setup, which the engine's template
copy sets explicitly.

Usage: python tools/xls2xlsx.py <in.xls> <out.xlsx>
"""
import sys
import datetime

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# xlrd border line-style index -> openpyxl style name
BORDER_STYLES = {
    1: "thin", 2: "medium", 3: "dashed", 4: "dotted", 5: "thick",
    6: "double", 7: "hair", 8: "mediumDashed", 9: "dashDot",
    10: "mediumDashDot", 11: "dashDotDot", 12: "mediumDashDotDot",
    13: "slantDashDot",
}
HALIGN = {1: "left", 2: "center", 3: "right", 4: "fill", 5: "justify"}
VALIGN = {0: "top", 1: "center", 2: "bottom", 3: "justify"}


def _font_rgb(book, colour_index):
    rgb = book.colour_map.get(colour_index)
    if not rgb or rgb == (0, 0, 0):
        return None
    return "FF%02X%02X%02X" % rgb


def convert(src, dst):
    book = xlrd.open_workbook(src, formatting_info=True)
    out = Workbook()
    out.remove(out.active)

    for sh in book.sheets():
        ws = out.create_sheet(title=sh.name)

        for c, info in (sh.colinfo_map or {}).items():
            if info.width:
                ws.column_dimensions[get_column_letter(c + 1)].width = info.width / 256.0
        for r, info in (sh.rowinfo_map or {}).items():
            if info.height:
                ws.row_dimensions[r + 1].height = info.height / 20.0

        for r in range(sh.nrows):
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                xf = book.xf_list[sh.cell_xf_index(r, c)]

                value = cell.value
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    value = None
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = datetime.datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))
                    except Exception:
                        pass
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(cell.value)
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    value = None
                elif cell.ctype == xlrd.XL_CELL_BLANK:
                    value = None

                fnt = book.font_list[xf.font_index]
                bl = xf.border
                edges = {
                    "left": bl.left_line_style, "right": bl.right_line_style,
                    "top": bl.top_line_style, "bottom": bl.bottom_line_style,
                }
                has_style = any(edges.values()) or fnt.bold or fnt.italic
                if value is None and not has_style:
                    continue

                target = ws.cell(row=r + 1, column=c + 1, value=value)
                target.font = Font(
                    name=fnt.name or "Arial",
                    size=(fnt.height or 200) / 20.0,
                    bold=bool(fnt.bold),
                    italic=bool(fnt.italic),
                    underline="single" if fnt.underline_type else None,
                    color=_font_rgb(book, fnt.colour_index),
                )
                # always pass all four Side objects (style=None when unset) so
                # the result matches native Excel files, where every border
                # side exists as a Side instance
                target.border = Border(**{
                    side: Side(style=BORDER_STYLES.get(style_ix) if style_ix else None)
                    for side, style_ix in edges.items()
                })
                al = xf.alignment
                target.alignment = Alignment(
                    horizontal=HALIGN.get(al.hor_align),
                    vertical=VALIGN.get(al.vert_align),
                    wrap_text=bool(al.text_wrapped) or None,
                )
                fmt = book.format_map.get(xf.format_key)
                if fmt and fmt.format_str and fmt.format_str != "General":
                    target.number_format = fmt.format_str

        for rlo, rhi, clo, chi in sh.merged_cells:
            ws.merge_cells(start_row=rlo + 1, end_row=rhi,
                           start_column=clo + 1, end_column=chi)

    out.save(dst)
    print(f"converted {src} -> {dst}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    convert(sys.argv[1], sys.argv[2])
